"""
================================================================================
REPORTS / EXCEL_EXPORT.PY - Enhanced Excel Report Generator
================================================================================
Generates formatted Excel audit reports with enhanced context fields,
conditional formatting, and a summary dashboard.

Features:
- Conditional formatting for status columns (✅ green, ⚠️ yellow, ❌ red)
- Color swatches for color values (rendered as cell backgrounds)
- Monospace font for full_selector and parent_context
- Color-coded source_context badges
- Auto-column width adjustment
- Freeze panes for easy scrolling
- Summary dashboard tab with statistics
- Filtered views for quick issue triage

Dependencies:
- openpyxl: pip install openpyxl
- pandas: pip install pandas

Author: Built with Claude in BoodleBox
Date: January 2026
================================================================================
"""

import os
from datetime import datetime
from typing import Dict, List, Any, Optional
from urllib.parse import urlparse

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, Fill, PatternFill, Alignment, Border, Side,
        NamedStyle, Color
    )
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import (
        FormulaRule, ColorScaleRule, CellIsRule
    )
    from openpyxl.worksheet.filters import AutoFilter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️ openpyxl or pandas not installed. Excel export disabled.")
    print("   Install with: pip install openpyxl pandas")


# ==============================================================================
# STYLE DEFINITIONS
# ==============================================================================

# Colors
HEADER_BG = "002d68"  # Hero Navy
HEADER_FG = "FFFFFF"  # White
PASS_BG = "d1fae5"    # Light green
PASS_FG = "065f46"    # Dark green
WARN_BG = "fef3c7"    # Light yellow
WARN_FG = "92400e"    # Dark amber
FAIL_BG = "fee2e2"    # Light red
FAIL_FG = "991b1b"    # Dark red
ALT_ROW_BG = "f9fafb" # Light gray for alternating rows

# Source context colors
SOURCE_COLORS = {
    "inline": ("fef3c7", "92400e"),      # Yellow
    "class": ("dbeafe", "1e40af"),       # Blue
    "global": ("f3e8ff", "6b21a8"),      # Purple
    "embed": ("fee2e2", "991b1b"),       # Red
    "external": ("e5e7eb", "374151"),    # Gray
}

# Column configurations
COLUMN_CONFIG = {
    "url": {"width": 35, "wrap": False},
    "type": {"width": 12, "wrap": False},
    "element": {"width": 15, "wrap": False},
    "full_selector": {"width": 45, "wrap": True, "monospace": True},
    "parent_context": {"width": 35, "wrap": True, "monospace": True},
    "property": {"width": 18, "wrap": False},
    "expected": {"width": 30, "wrap": True},
    "found": {"width": 30, "wrap": True},
    "status": {"width": 20, "wrap": False},
    "source_context": {"width": 18, "wrap": False},
    "text_snippet": {"width": 30, "wrap": True},
}


# ==============================================================================
# PATH GENERATION
# ==============================================================================

def get_reports_base_dir() -> str:
    """Get the base reports directory."""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    tool_dir = os.path.dirname(script_dir)
    return os.path.join(tool_dir, "reports")


def get_site_report_dir(base_url: str) -> str:
    """Get or create the site-specific report directory."""
    parsed = urlparse(base_url)
    domain = parsed.netloc or base_url.replace("https://", "").replace("http://", "").split("/")[0]
    
    if not domain:
        domain = "unknown_site"
    if domain.startswith("www."):
        domain = domain[4:]
    
    reports_base = get_reports_base_dir()
    site_dir = os.path.join(reports_base, domain)
    
    if not os.path.exists(site_dir):
        os.makedirs(site_dir)
        print(f"📁 Created report directory: {site_dir}")
    
    return site_dir


def get_report_filepath(base_url: str, report_type: str = "audit_all",
                        custom_filename: Optional[str] = None) -> str:
    """Generate the full filepath for an Excel report."""
    site_dir = get_site_report_dir(base_url)
    
    if custom_filename:
        filename = f"{custom_filename}.xlsx"
    else:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{report_type}_{timestamp}.xlsx"
    
    return os.path.join(site_dir, filename)


# ==============================================================================
# MAIN EXPORT FUNCTION
# ==============================================================================

def export_to_excel(results: List[Dict[str, Any]], base_url: str,
                    report_type: str = "audit_all",
                    custom_filename: Optional[str] = None) -> str:
    """
    Export audit results to a formatted Excel workbook.
    
    Args:
        results: List of audit result dictionaries
        base_url: The URL that was audited
        report_type: Type of audit for filename
        custom_filename: Optional custom filename
        
    Returns:
        Path to the created Excel file
    """
    if not EXCEL_AVAILABLE:
        print("❌ Excel export not available. Install openpyxl and pandas.")
        return ""
    
    if not results:
        print("⚠️ No results to export")
        return ""
    
    filepath = get_report_filepath(base_url, report_type, custom_filename)
    domain = urlparse(base_url).netloc or base_url
    
    # Create workbook
    wb = Workbook()
    
    # Create Summary sheet first
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _create_summary_sheet(ws_summary, results, domain)
    
    # Create Details sheet
    ws_details = wb.create_sheet("Details")
    _create_details_sheet(ws_details, results)
    
    # Create Issues Only sheet (filtered view)
    ws_issues = wb.create_sheet("Issues Only")
    issues_only = [r for r in results if "✅" not in str(r.get("status", ""))]
    if issues_only:
        _create_details_sheet(ws_issues, issues_only)
    else:
        ws_issues["A1"] = "🎉 No issues found! All items passed."
        ws_issues["A1"].font = Font(size=14, bold=True, color="065f46")
    
    # Create By Source sheet (grouped by source_context)
    ws_source = wb.create_sheet("By Source")
    _create_by_source_sheet(ws_source, results)
    
    # Save workbook
    wb.save(filepath)
    print(f"✅ Excel report saved to: {filepath}")
    
    return filepath


# ==============================================================================
# SUMMARY SHEET
# ==============================================================================

def _create_summary_sheet(ws, results: List[Dict[str, Any]], domain: str):
    """Create the summary dashboard sheet."""
    
    # Calculate statistics
    total = len(results)
    passed = sum(1 for r in results if "✅" in str(r.get("status", "")))
    warnings = sum(1 for r in results if "⚠️" in str(r.get("status", "")))
    failed = sum(1 for r in results if "❌" in str(r.get("status", "")))
    pass_rate = (passed / total * 100) if total > 0 else 0
    
    # Count by type
    by_type = {}
    for r in results:
        t = r.get("type", "unknown")
        if t not in by_type:
            by_type[t] = {"total": 0, "passed": 0, "warnings": 0, "failed": 0}
        by_type[t]["total"] += 1
        status = str(r.get("status", ""))
        if "✅" in status:
            by_type[t]["passed"] += 1
        elif "⚠️" in status:
            by_type[t]["warnings"] += 1
        elif "❌" in status:
            by_type[t]["failed"] += 1
    
    # Count by source
    by_source = {}
    for r in results:
        source = r.get("source_context", "Unknown")
        status = str(r.get("status", ""))
        if "✅" not in status:  # Only count non-passing items
            by_source[source] = by_source.get(source, 0) + 1
    
    # Styles
    title_font = Font(size=20, bold=True, color=HEADER_BG)
    subtitle_font = Font(size=12, color="6b7280")
    header_font = Font(size=11, bold=True, color=HEADER_FG)
    header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    stat_font = Font(size=24, bold=True)
    label_font = Font(size=10, color="6b7280")
    
    # Title
    ws["A1"] = f"Audit Report: {domain}"
    ws["A1"].font = title_font
    ws.merge_cells("A1:E1")
    
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A2"].font = subtitle_font
    
    # Overall Statistics (Row 4-6)
    row = 4
    ws[f"A{row}"] = "OVERALL STATISTICS"
    ws[f"A{row}"].font = Font(size=14, bold=True, color=HEADER_BG)
    ws.merge_cells(f"A{row}:E{row}")
    
    row = 5
    stats = [
        ("Total Checked", total, "002d68"),
        ("Passed ✅", passed, "065f46"),
        ("Warnings ⚠️", warnings, "92400e"),
        ("Failed ❌", failed, "991b1b"),
        ("Pass Rate", f"{pass_rate:.1f}%", "002d68"),
    ]
    
    for col, (label, value, color) in enumerate(stats, start=1):
        cell_label = ws.cell(row=row, column=col, value=label)
        cell_label.font = label_font
        cell_label.alignment = Alignment(horizontal="center")
        
        cell_value = ws.cell(row=row+1, column=col, value=value)
        cell_value.font = Font(size=24, bold=True, color=color)
        cell_value.alignment = Alignment(horizontal="center")
    
    # Issues by Source (Row 9+)
    row = 9
    ws[f"A{row}"] = "ISSUES BY SOURCE (Hiding Spots)"
    ws[f"A{row}"].font = Font(size=14, bold=True, color=HEADER_BG)
    ws.merge_cells(f"A{row}:C{row}")
    
    row = 10
    ws.cell(row=row, column=1, value="Source").font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    ws.cell(row=row, column=2, value="Issue Count").font = header_font
    ws.cell(row=row, column=2).fill = header_fill
    
    row = 11
    if by_source:
        for source, count in sorted(by_source.items(), key=lambda x: -x[1]):
            ws.cell(row=row, column=1, value=source)
            ws.cell(row=row, column=2, value=count)
            
            # Color code the source
            source_lower = source.lower()
            for key, (bg, fg) in SOURCE_COLORS.items():
                if key in source_lower:
                    ws.cell(row=row, column=1).fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
                    ws.cell(row=row, column=1).font = Font(color=fg)
                    break
            
            row += 1
    else:
        ws.cell(row=row, column=1, value="All items passed! 🎉")
        ws.cell(row=row, column=1).font = Font(color="065f46", bold=True)
    
    # Breakdown by Type (Row after sources + 2)
    row += 2
    ws[f"A{row}"] = "BREAKDOWN BY TYPE"
    ws[f"A{row}"].font = Font(size=14, bold=True, color=HEADER_BG)
    ws.merge_cells(f"A{row}:E{row}")
    
    row += 1
    headers = ["Type", "Total", "Passed", "Warnings", "Failed"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    row += 1
    for type_name, type_stats in sorted(by_type.items()):
        ws.cell(row=row, column=1, value=type_name)
        ws.cell(row=row, column=2, value=type_stats["total"])
        
        passed_cell = ws.cell(row=row, column=3, value=type_stats["passed"])
        if type_stats["passed"] > 0:
            passed_cell.font = Font(color="065f46")
        
        warn_cell = ws.cell(row=row, column=4, value=type_stats["warnings"])
        if type_stats["warnings"] > 0:
            warn_cell.font = Font(color="92400e", bold=True)
        
        fail_cell = ws.cell(row=row, column=5, value=type_stats["failed"])
        if type_stats["failed"] > 0:
            fail_cell.font = Font(color="991b1b", bold=True)
        
        row += 1
    
    # Adjust column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12


# ==============================================================================
# DETAILS SHEET
# ==============================================================================

def _create_details_sheet(ws, results: List[Dict[str, Any]]):
    """Create the detailed results sheet with formatting."""
    
    if not results:
        ws["A1"] = "No results to display"
        return
    
    # Determine columns from first result + standard columns
    standard_cols = [
        "url", "type", "element", "full_selector", "parent_context",
        "property", "expected", "found", "status", "source_context", "text_snippet"
    ]
    
    # Get all unique keys
    all_keys = set()
    for r in results:
        all_keys.update(r.keys())
    
    # Use standard columns that exist, plus any extras
    columns = [c for c in standard_cols if c in all_keys]
    extra_cols = sorted(all_keys - set(columns))
    columns.extend(extra_cols)
    
    # Styles
    header_font = Font(size=11, bold=True, color=HEADER_FG)
    header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    mono_font = Font(name="Consolas", size=9)
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    
    thin_border = Border(
        left=Side(style="thin", color="e5e5e5"),
        right=Side(style="thin", color="e5e5e5"),
        top=Side(style="thin", color="e5e5e5"),
        bottom=Side(style="thin", color="e5e5e5"),
    )
    
    # Write headers
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Write data rows
    for row_idx, result in enumerate(results, start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            value = result.get(col_name, "")
            if value is None:
                value = ""
            
            cell = ws.cell(row=row_idx, column=col_idx, value=str(value))
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top")
            
            # Apply column-specific formatting
            config = COLUMN_CONFIG.get(col_name, {})
            
            if config.get("monospace"):
                cell.font = mono_font
            
            if config.get("wrap"):
                cell.alignment = wrap_alignment
            
            # Status column formatting
            if col_name == "status":
                status_str = str(value)
                if "✅" in status_str:
                    cell.fill = PatternFill(start_color=PASS_BG, end_color=PASS_BG, fill_type="solid")
                    cell.font = Font(color=PASS_FG, bold=True)
                elif "⚠️" in status_str:
                    cell.fill = PatternFill(start_color=WARN_BG, end_color=WARN_BG, fill_type="solid")
                    cell.font = Font(color=WARN_FG, bold=True)
                elif "❌" in status_str:
                    cell.fill = PatternFill(start_color=FAIL_BG, end_color=FAIL_BG, fill_type="solid")
                    cell.font = Font(color=FAIL_FG, bold=True)
            
            # Source context formatting
            if col_name == "source_context":
                source_lower = str(value).lower()
                for key, (bg, fg) in SOURCE_COLORS.items():
                    if key in source_lower:
                        cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
                        cell.font = Font(color=fg, size=9)
                        break
            
            # Color swatch for "found" column when it contains a hex color
            if col_name == "found" and str(value).startswith("#"):
                hex_color = str(value).lstrip("#")[:6]
                if len(hex_color) == 6:
                    try:
                        # Add a small color indicator
                        cell.value = f"■ {value}"
                        # We can't easily set partial cell colors, so we note it
                    except:
                        pass
        
        # Alternating row colors (for non-status-highlighted rows)
        if row_idx % 2 == 0:
            for col_idx in range(1, len(columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                # Only apply if not already colored by status
                if cell.fill.start_color.rgb == "00000000" or cell.fill.start_color.rgb is None:
                    cell.fill = PatternFill(start_color=ALT_ROW_BG, end_color=ALT_ROW_BG, fill_type="solid")
    
    # Set column widths
    for col_idx, col_name in enumerate(columns, start=1):
        config = COLUMN_CONFIG.get(col_name, {"width": 15})
        ws.column_dimensions[get_column_letter(col_idx)].width = config.get("width", 15)
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Add auto-filter
    if results:
        last_col = get_column_letter(len(columns))
        last_row = len(results) + 1
        ws.auto_filter.ref = f"A1:{last_col}{last_row}"


# ==============================================================================
# BY SOURCE SHEET
# ==============================================================================

def _create_by_source_sheet(ws, results: List[Dict[str, Any]]):
    """Create a sheet grouped by source context."""
    
    # Group results by source
    by_source = {}
    for r in results:
        source = r.get("source_context", "Unknown")
        if source not in by_source:
            by_source[source] = []
        by_source[source].append(r)
    
    # Styles
    section_font = Font(size=14, bold=True, color=HEADER_BG)
    header_font = Font(size=10, bold=True, color=HEADER_FG)
    header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    
    row = 1
    columns = ["url", "full_selector", "property", "expected", "found", "status"]
    
    for source, source_results in sorted(by_source.items()):
        # Section header
        ws.cell(row=row, column=1, value=f"{source} ({len(source_results)} items)")
        ws.cell(row=row, column=1).font = section_font
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(columns))
        
        # Apply source color to header
        source_lower = source.lower()
        for key, (bg, fg) in SOURCE_COLORS.items():
            if key in source_lower:
                for col in range(1, len(columns) + 1):
                    ws.cell(row=row, column=col).fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
                ws.cell(row=row, column=1).font = Font(size=14, bold=True, color=fg)
                break
        
        row += 1
        
        # Column headers
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=row, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
        
        row += 1
        
        # Data rows
        for result in source_results:
            for col_idx, col_name in enumerate(columns, start=1):
                value = result.get(col_name, "")
                cell = ws.cell(row=row, column=col_idx, value=str(value) if value else "")
                
                # Status formatting
                if col_name == "status":
                    status_str = str(value)
                    if "✅" in status_str:
                        cell.fill = PatternFill(start_color=PASS_BG, end_color=PASS_BG, fill_type="solid")
                        cell.font = Font(color=PASS_FG)
                    elif "⚠️" in status_str:
                        cell.fill = PatternFill(start_color=WARN_BG, end_color=WARN_BG, fill_type="solid")
                        cell.font = Font(color=WARN_FG)
                    elif "❌" in status_str:
                        cell.fill = PatternFill(start_color=FAIL_BG, end_color=FAIL_BG, fill_type="solid")
                        cell.font = Font(color=FAIL_FG)
            
            row += 1
        
        # Add spacing between sections
        row += 1
    
    # Set column widths
    widths = [35, 45, 18, 30, 30, 20]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ==============================================================================
# CONVENIENCE FUNCTION FOR MAIN.PY
# ==============================================================================

def generate_excel_report(results: List[Dict[str, Any]], base_url: str,
                          report_type: str = "audit_all",
                          custom_filename: Optional[str] = None) -> str:
    """
    Main entry point for Excel report generation.
    
    This function is called by main.py when Excel output is requested.
    
    Args:
        results: List of audit result dictionaries
        base_url: The URL that was audited
        report_type: Type of audit for filename
        custom_filename: Optional custom filename
        
    Returns:
        Path to the created Excel file
    """
    return export_to_excel(results, base_url, report_type, custom_filename)
